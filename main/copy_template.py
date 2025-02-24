import os
import shutil
from openpyxl import load_workbook
import statistics


# Directory paths
RAW_DIR = "../raw"
RESULT_DIR = "../result"
TEMPLATE_FILE = "../template2.xlsx"

def extract_product_info(sheet):
    """Extract product information from the '产品' sheet."""
    general_info = sheet["A1"].value.split("\n")
    info_dict = {}

    for line in general_info:
        # DEBUG
        print("the line: " + line)
        if "：" in line:  # Ensure it has the delimiter
            key, value = line.split("：", 1)  # Split at the first colon
            info_dict[key.strip()] = value.strip()

    return {
        "node_id": info_dict.get("类目ID（NodeId）"),
        "name": info_dict.get("类目"),
        "path": info_dict.get("路径"),
        "url": info_dict.get("BSR URL"),
    }

def calculate_percentage(col, sheet):
    """Calculate the percentage of (sum of the top 10 sorted values) / (sum of all values) in a specified column."""
    values = []
    for i in range(3, 103):
        value = sheet[f"{col}{i}"].value
        try:
            # Try converting to float first (in case of decimals)
            num_value = float(value)
        except TypeError:
            # If value is None, set it to 0
            num_value = 0
        except ValueError:
            # If the value can't be converted to a float, also set it to 0
            num_value = 0
        values.append(num_value)

    sorted_values = sorted(values, reverse=True)
    # Calculate sums of the top 10 and all sorted values
    sum_top_10 = sum(sorted_values[:10])  # Sum of the first 10 values after sorting
    sum_all = sum(sorted_values)          # Sum of all values

    # DEBUG
    print(f"Sum of top 10 values: {sum_top_10}")
    print(f"Sum of all values: {sum_all}")

    if sum_all == 0:
        return "0.0%"  # Avoid division by zero

    # Calculate and return the percentage
    percentage = (sum_top_10 / sum_all) * 100
    return f"{percentage:.1f}%"

def new_product_analysis(days, sheet):
    """Convert integers below 180 in column 'L' to 180, count conversions, and calculate selling percentage."""
    print("new_product_analysis")
    total_sum_i = 0
    modified_sum_i = 0
    count = 0

    for i in range(3, sheet.max_row + 1):
        cell_l = sheet[f"L{i}"]
        cell_i = sheet[f"I{i}"]

        # Ensure that cell_i can be converted to float for summation
        try:
            value_i = float(cell_i.value)
        except (ValueError, TypeError):
            value_i = 0

        # Add to total sum of column 'I'
        total_sum_i += value_i

        try:
            # Force conversion of the cell value to an integer
            days_value = int(cell_l.value)
            if days_value < days:
                count += 1  # Increment the count for each conversion
                modified_sum_i += value_i  # Add the value from column 'I' for this row
        except TypeError:
            continue  # Skip cells with non-integer types

    # Calculate the percentage
    if total_sum_i > 0:
        percentage = (modified_sum_i / total_sum_i) * 100
    else:
        percentage = 0  # Avoid division by zero

    return count, percentage

def calculate_median(sheet, column):
    """Calculate the median of integers in the specified column starting from row 3."""
    values = []
    for row in range(3, sheet.max_row + 1):  # Start from row 3 to the last row
        cell_value = sheet[f"{column}{row}"].value
        try:
            # Convert the value to float and append to the list
            f_value = float(cell_value)
            values.append(f_value)
        except (TypeError, ValueError):
            # Ignore cells with non-convertible values
            continue

    if values:
        return statistics.median(values)
    else:
        return None  # Return None if there are no valid integer values

def calculate_last_x_average(x, sheet, column, exclute_zero):
    """Calculate the average of the last 25% of rows for the given column after sorting values in descending order."""
    values = []
    # Collect all values from the column
    for row in range(3, sheet.max_row + 1):  # Assuming data starts from row 3
        cell = sheet[f"{column}{row}"].value
        try:
            # Convert the cell value to float
            float_value = float(cell)
            if(exclute_zero and float_value == 0):
                continue
            values.append(float_value)
        except (TypeError, ValueError):
            # Ignore cells with non-convertible values
            continue

    if not values:
        return None  # Return None if there are no valid float values

    # Sort values in descending order
    values.sort(reverse=True)

    # Calculate the index for the last x% of the sorted list
    start_index = int(len(values) * (1 - x*0.01))
    last_quarter_values = values[start_index:]

    # Return the average of the last quarter of the list
    if last_quarter_values:
        return sum(last_quarter_values) / len(last_quarter_values)
    else:
        return None


def process_file(file_path):
    """Process a single raw data file."""
    # Load the raw data workbook and access the '产品' sheet
    raw_wb = load_workbook(file_path)
    product_sheet = raw_wb["产品"]
    cal_sheet = raw_wb["产品详情"]

    # Extract product information
    product_info = extract_product_info(product_sheet)
    this_name = product_info["name"]
    
    if not this_name:
        print(f"Skipping {file_path}: Missing product name.")
        return

    # Create a copy of the template
    new_file_name = f"{this_name}_template.xlsx"
    new_file_path = os.path.join(RESULT_DIR, new_file_name)
    shutil.copyfile(TEMPLATE_FILE, new_file_path)

    # Load the copied template
    template_wb = load_workbook(new_file_path)
    template_sheet = template_wb.active  # Assumes writing to the first sheet

    # Copy general information to the result sheet
    template_sheet["B4"] = this_name
    template_sheet["B5"] = product_info["url"]
    template_sheet["B6"] = product_info["path"]

    # New product analysis
    num_np, per_np = new_product_analysis(181, cal_sheet)

    # Copy specific cells from raw to template
    template_sheet["B11"] = product_sheet["B2"].value
    template_sheet["B12"] = product_sheet["D2"].value
    template_sheet["B13"] = product_sheet["F2"].value
    template_sheet["B14"] = num_np
    # template_sheet["B15"] = product_sheet["H4"].value
    template_sheet["B16"] = product_sheet["J2"].value
    template_sheet["B17"] = calculate_last_x_average(25, cal_sheet, "H", False)

    template_sheet["D11"] = product_sheet["B3"].value
    template_sheet["D12"] = product_sheet["F4"].value
    template_sheet["D13"] = product_sheet["F3"].value
    template_sheet["D14"] = per_np
    # template_sheet["D15"] = product_sheet["H5"].value
    template_sheet["D16"] = product_sheet["J3"].value
    template_sheet['D17'].number_format = '0.00'
    template_sheet["D17"] = calculate_last_x_average(25, cal_sheet, "I", True)

    template_sheet["F11"] = product_sheet["B4"].value
    template_sheet["F12"] = f"{product_sheet['D4'].value}/{product_sheet['D5'].value}"
    template_sheet["F13"] = product_sheet["H2"].value
    template_sheet["F14"] = product_sheet["H3"].value
    template_sheet["F15"] = calculate_median(cal_sheet, "L")
    template_sheet["F16"] = product_sheet["L4"].value
    
    # Calculate percentage and write to D16
    percentage1 = calculate_percentage("I", cal_sheet)
    template_sheet["F17"] = percentage1
    
    # percentage2 = calculate_percentage("H", cal_sheet)
    # template_sheet["F16"] = percentage2

    # Save the modified template
    template_wb.save(new_file_path)
    print(f"Processed and saved: {new_file_path}")

def main():
    """Main function to iterate through raw files and process them."""
    for file_name in os.listdir(RAW_DIR):
        # DEBUG 
        print("file name: " + file_name)
        if file_name.startswith("~$"):
            continue
        if file_name.endswith("产品列表.xlsx"):
            file_path = os.path.join(RAW_DIR, file_name)
            print(f"Processing file: {file_name}")
            process_file(file_path)

if __name__ == "__main__":
    main()
