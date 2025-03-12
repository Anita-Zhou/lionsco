import os
import shutil
import pandas as pd
from openpyxl import load_workbook
import statistics
import re

# Directory paths
RAW_DIR = "../raw"
RESULT_DIR = "../result"
TEMPLATE_FILE = "../template3.xlsx"

产品名 = 'A3'
类目链接 = 'D3'
在售产品数 = 'F3'
买家垄断系数 = 'F4'
月均销额 = 'F5'
AMZ自营占比 = 'F6'
平均上架天数 = 'F7'
新品占比 = 'F8'
月搜索量 = 'F9'
后25销额 = 'F10'
在售商家数 = 'F11'
平均价格 = 'H3'
月均销量 = 'H4'
中位销额 = 'H5'
AMZ自营额 = 'H6'
平均评价数量 = 'H7'
新品市场份额 = 'H8'
退货率 = 'H9'
平均FBA占比 = 'H10'
品牌数 = 'H11'

this_name = 'TEST'
top_num = None
wanted_keyword = None
unwanted_keyword = None
days = None

def extract_product_name(filename):
    """从文件名中提取产品名称"""
    match = re.search(r'_(.*?)_([^_]+)产品看板导出', filename)
    return match.group(2) if match else None

# Example implementations of previous functions using pandas
def cal_average(df, column_name):
    # Temporarily drop rows where this column is 0
    valid_values = df[column_name][df[column_name] != 0]  
    # Calculate mean only using valid values
    return valid_values.mean() if not valid_values.empty else 0

def cal_median(df, column_name):
    # Temporarily drop rows where this column is 0
    valid_values = df[column_name][df[column_name] != 0]  
    # # DEBUG
    # print("cal_median")
    # print(valid_values)
    return valid_values.median() if not valid_values.empty else 0

def cal_last_x_percent(df, column_name, x=25):
    # Temporarily drop rows where this column is 0
    valid_values = df[column_name][df[column_name] != 0]  
    if valid_values.empty:
        return 0
    sorted_values = valid_values.sort_values(ascending=False)
    n = max(1, int(len(sorted_values) * (x / 100)))
    return sorted_values.tail(n).mean() if n > 0 else 0

def count_keyword(df, column_name, keyword):
    # # DEBUG
    # print("keyword: " + keyword)
    valid_df = df[df[column_name] == keyword]
    # print(valid_df)
    # print(valid_df.shape[0])
    return valid_df.shape[0]

def count_new_product(df, column_name, days):
    # Temporarily drop rows where this column is 0
    valid_df = df[df[column_name] != 0]  
    return (valid_df[column_name] < days).sum() if not valid_df.empty else 0

def sum_top_x(df, column_name, x=10):
    return df[column_name].nlargest(x).sum() if not df[column_name].empty else 0

def count_unique(df, column_name):
    return df[column_name].nunique()

def create_avg_fba(df):
    # Convert columns to numeric, forcing non-numeric values to NaN
    df['FBA费用($)'] = pd.to_numeric(df['FBA费用($)'], errors='coerce')
    df['实际价格($)'] = pd.to_numeric(df['实际价格($)'], errors='coerce')
    # Calculate the ratio, handling NaN by replacing them with 0
    df['FBA费用占比'] = (df['FBA费用($)'] / df['实际价格($)']).fillna(0)
    # Replace infinity values (from division by zero) with 0
    df['FBA费用占比'] = df['FBA费用占比'].replace([float('inf'), -float('inf')], 0)


"""
Start processing a single raw data file using pandas.
"""
def process_file(file_path):
    """Process a single raw data file using pandas."""
    # Load specific sheets from the Excel file
    # Assume header row is row 3
    product_df = pd.read_excel(file_path, sheet_name="产品", header=2) 
    total_num = product_df.shape[0]
    # print(total_num)
    print("--------------")
    # ============================
    # Step 1: Get User Preferences
    # ============================
    global this_name, top_num, wanted_keyword, unwanted_keyword  # Declare global variables to modify them

    # Prompt user whether they want top 400 data
    top_num = input(f"共有{total_num}个产品， 你想要看头部多少个产品的数据？(填入数字，或按 Enter 键跳过): ").strip()
    top_num = int(top_num) if top_num.isdigit() else None  # Convert to int if possible

    # Prompt user whether they want top 400 data
    days = input(f"你想要看上架天数多少天以内的产品数据？(填入数字，或按 Enter 键跳过): ").strip()
    days = int(days) if days.isdigit() else None  # Convert to int if possible

    # Prompt user for a keyword filter (optional)
    wanted_keyword = input("输入你想要的关键词 (或按 Enter 键跳过): ").strip()
    wanted_keyword = wanted_keyword if wanted_keyword else None  # Convert empty input to None

    # Prompt user for a keyword filter (optional)
    unwanted_keyword = input("输入你不想要的关键词 (或按 Enter 键跳过): ").strip()
    unwanted_keyword = unwanted_keyword if unwanted_keyword else None  # Convert empty input to None



    print("\n**** User Selections ****")
    print(f"Top {top_num} 产品数据")
    print(f"Wanted Keyword: {wanted_keyword if wanted_keyword else 'No wanted keyword filter'}")
    print(f"Unwanted Keyword: {unwanted_keyword if unwanted_keyword else 'No unwanted keyword filter'}")
    print("************************\n")
    
    # # DEBUG
    # # Print all column names to check for typos or formatting issues
    # print("Column Names:", product_df.columns.tolist())

    # # DEBUG
    # # Identify rows where "Listing月销量" is not numeric
    # non_numeric_rows = product_df[~product_df["Listing月销量"].astype(str).str.replace('.', '', regex=False).str.isdigit()]
    # # Print the problematic rows
    # print("Rows with non-numeric 'Listing月销量':")
    # print(non_numeric_rows)

    # ================================
    #   Step 2: Filter out NA values 
    # ================================
    cols = ['Listing月销量', 'Listing月销额($)', '实际价格($)', '上架时长(天）']
    for col in cols:
        if col in product_df.columns:
            # Convert to numeric, replacing non-numeric values with 0
            product_df[col] = pd.to_numeric(product_df[col], errors='coerce').fillna(0)
        else:
            print(f"Warning: Column '{col}' not found in DataFrame")
    #===================
    # Drop unwanted rows
    #===================
    # Filter to keep rows where "产品名称" or "五点描述" contains <wanted_keyword> (case-insensitive)
    if wanted_keyword:  # This ensures None or empty strings are ignored
        product_df = product_df[
            product_df["产品名称"].astype(str).str.lower().str.contains(wanted_keyword, na=False) |
            product_df["五点描述"].astype(str).str.lower().str.contains(wanted_keyword, na=False)
        ].dropna(subset=["产品名称", "五点描述"], how="all")  # Drop rows where both columns are NaN
        this_name += f"_w_{wanted_keyword}"

    # Filter to leave rows where "产品名称" or "五点描述" contains <unwanted_keyword> (case-insensitive)
    if unwanted_keyword:  # Ensure unwanted_keyword is not None or empty
        product_df = product_df[
            ~product_df["产品名称"].astype(str).str.lower().str.contains(unwanted_keyword, na=False) &
            ~product_df["五点描述"].astype(str).str.lower().str.contains(unwanted_keyword, na=False)
        ]

    if top_num and (top_num <= total_num):
        # Drop rows with value larger than 400
        product_df = product_df[product_df['序号'] <= top_num]
        this_name += f"_top{top_num}"

    if days:
        # Drop rows with value larger than <days>
        product_df = product_df[product_df['上架时长(天）'] <= days]
        this_name += f"_days{days}"

    # Drop rows with value smaller than 10
    product_df = product_df[product_df['Listing月销量'] >= 10]


    create_avg_fba(product_df)
    # # DEBUG
    # # Print all column names to check for typos or formatting issues
    # print("Column Names:", product_df.columns.tolist())

    # =============================================================== #
    # ================  Step 2: Copy filtered raw data ============== #
    # Create a copy of the initially filtered data
    required_columns = [
        '序号', '主图', '产品名称', 'ASIN', 'URL', 'Listing月销量', 'Listing月销额($)', 
        '五点描述', '实际价格($)', '品牌', 'BBX卖家属性', '店铺', '国籍/地区', 
        '上架时长(天）', '广告花费指数', '评价数量', 'FBA费用($)', '变体数量', '是否抛货', 
        'Color', 'Material', 'Size', 'FBA费用占比'
    ]

    # Ensure only existing columns are selected (avoid errors if some columns are missing)
    available_columns = [col for col in required_columns if col in product_df.columns]
    raw_data = product_df[available_columns].copy()  # Save a copy of the raw data for reference

    # =============================================================== #
    # ================  Step 3: Caulate integrated data ============= #
    # You can now use these functions with column names instead of Excel coordinates
    # Example usage (you'll need to map your Excel columns to DataFrame column names):
    results = {
        'avg_sales': cal_average(product_df, 'Listing月销额($)'), 
        'median_sales': cal_median(product_df, 'Listing月销额($)'), 
        'total_prdcts': count_unique(product_df, 'ASIN'),    
        'mkt_barrier' :( sum_top_x(product_df, 'Listing月销额($)', 10))/product_df['Listing月销额($)'].sum(),
        'last25_sales': cal_last_x_percent(product_df, 'Listing月销额($)', 25),
        'unique_brands': count_unique(product_df, '品牌'),
        'unique_sellers': count_unique(product_df, '店铺'),
        'num_amz': count_keyword(product_df, 'BBX卖家属性', '亚马逊自营'),
        'avg_days': cal_average(product_df, '上架时长(天）'),
        'new_prdcts': count_new_product(product_df, '上架时长(天）', 181),
        'avg_price': cal_average(product_df, '实际价格($)'),
        'avg_num_sales': cal_average(product_df, 'Listing月销量'),
        'avg_reviews': cal_average(product_df, '评价数量'),
        'amz_share': (product_df.loc[product_df['BBX卖家属性'] == '亚马逊自营', 'Listing月销额($)'].sum())/(product_df['Listing月销额($)'].sum()),
        'new_share': (product_df.loc[product_df['上架时长(天）'] < 181, 'Listing月销额($)'].sum())/(product_df['Listing月销额($)'].sum()),
        'avg_fba': cal_average(product_df, 'FBA费用占比')
    }
    
    return results, raw_data

'''
Copy the given template and save the processed results to a new Excel file.
'''
def save_result(result, raw):
    """Save the processed results to a new Excel file based on the template."""
    # DEBUG
    print("Saving...")
    # Create a copy of the template
    new_file_name = f"{this_name}_template.xlsx"
    new_file_path = os.path.join(RESULT_DIR, new_file_name)
    shutil.copyfile(TEMPLATE_FILE, new_file_path)

    # Load the copied template
    template_wb = load_workbook(new_file_path)
    template_sheet = template_wb.active  # Assumes writing to the first sheet

    # =========================================== #
    # ================  Step 1  ================= #
    # Copy specific cells from raw to template
    template_sheet[在售产品数] = result['total_prdcts']
    template_sheet[买家垄断系数] = result['mkt_barrier']
    template_sheet[月均销额] = result['avg_sales']
    template_sheet[AMZ自营占比] = str(result['num_amz']) + "个, " + str(result['num_amz'] / result['total_prdcts']*100) + "%"
    template_sheet[平均上架天数] = result['avg_days']
    template_sheet[新品占比] = str(result['new_prdcts']) + "个, " + str(result['new_prdcts'] / result['total_prdcts']*100) + "%"
    template_sheet[月搜索量]
    template_sheet[后25销额] = result['last25_sales']
    template_sheet[在售商家数] = result['unique_sellers']
    template_sheet[平均价格] = result['avg_price']
    template_sheet[月均销量] = result['avg_num_sales']
    template_sheet[中位销额] = result['median_sales']
    template_sheet[AMZ自营额] = result['amz_share']
    template_sheet[平均评价数量] = result['avg_reviews']
    template_sheet[新品市场份额] = result['new_share']
    template_sheet[退货率]
    template_sheet[平均FBA占比] = result['avg_fba']
    template_sheet[品牌数] = result['unique_brands']

    # =========================================== #
    # ================  Step 2  ================= #
    # Create a new sheet for filtered data
    extra_sheet = template_wb.create_sheet(title="产品") #Create a second sheet
    # Write DataFrame headers to Excel
    for col_idx, col_name in enumerate(raw.columns, start=1):
        extra_sheet.cell(row=1, column=col_idx, value=col_name)

    # Write DataFrame values to Excel
    for row_idx, row in enumerate(raw.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            extra_sheet.cell(row=row_idx, column=col_idx, value=value)

    # Save the modified template
    template_wb.save(new_file_path)
    print(f"Processed and saved: {new_file_path}")


'''
Execution starts here.
'''
def main():
    """Main function to iterate through raw files and process them."""
    # ============================
    # Step 1: Process Raw Files
    # ============================
    for file_name in os.listdir(RAW_DIR):
        # # DEBUG 
        # print("file name: " + file_name)
        if file_name.startswith("~$"):
            continue
        if file_name.endswith("产品看板导出.xlsx"):
            print("=====================================")
            file_path = os.path.join(RAW_DIR, file_name)
            print(f"Processing file: {file_name}")
            global this_name
            this_name = extract_product_name(file_name)
            print(f"Product Name: {this_name}")
            results, raw_data = process_file(file_path)
            # Do something with results, like write to template
            save_result(results, raw_data)
            print("=====================================\n")

if __name__ == "__main__":
    main()