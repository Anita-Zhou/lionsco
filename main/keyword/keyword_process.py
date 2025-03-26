import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from collections import Counter

# Directory paths
RAW_DIR = "../../raw/keyword"
RESULT_DIR = "../../result/keyword"

# Prompt user whether they want top 400 data
dir_name = input(f"请输入你想要处理的文件夹名称 (注：请先将需要处理的文件合并在一个文件夹中): ").strip()
input_folder = f"{RAW_DIR}/{dir_name}"
output_file = os.path.join(RESULT_DIR, dir_name + "_combined.xlsx")

def count_word_freq(df, col_name):
    all_words = []
    for phrase in df[col_name].dropna():
        words = phrase.strip().split()
        all_words.extend(words)
    return Counter(all_words)
    
def post_process(filename):
    #=====================
    #   Post-processing
    #=====================
    print(f"Post-processing file: {filename}")
    wb = load_workbook(filename)
    for sheetname in wb.sheetnames:
    
        ws = wb[sheetname]

        # Set row height of new first row (was originally third)
        ws.row_dimensions[1].height = 30

        # Fill background color for new header row
        fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        for cell in ws[1]:
            cell.fill = fill
        
        # Set each column's width to 13 characters
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 13

        # Save the styled workbook
        wb.save(filename)
        print(f"Styled Excel saved to: {filename}")
        print("=====================================")
        print("                 END                 ")
        print("=====================================")

# # Initialize an Excel writer
# writer = pd.ExcelWriter(output_file, engine='openpyxl')

# Loop through all Excel files in the folder
for filename in os.listdir(input_folder): 
    if filename.startswith("~$"):
        continue
    if filename.endswith(".xlsx") and "_反查出单词列表" in filename:
        print("=====================================")
        print("                 START               ")
        print("=====================================")

        # Initialize an Excel writer
        filepath = os.path.join(input_folder, filename)
        out = os.path.join(RESULT_DIR, filename)
        print("Output file: ", out)
        writer = pd.ExcelWriter(out, engine='openpyxl')
        print(f"{filepath}")

        # Extract ASIN from filename (assumes format: <date>_<ASIN>_反查出单词列表.xlsx)
        try:
            asin = filename.split("_")[1]
            print(f"Processing file: {asin}")
        except IndexError:
            print(f"Skipping {filename}: ASIN not found in filename.")
            continue

        # Read the '产品' sheet
        try:
            df = pd.read_excel(filepath, sheet_name="产品", header=2)
        except Exception as e:
            print(f"Error reading file {filename}: {e}")
            continue

        # ====================
        #      Do process  
        # ====================
        # Replace all NA with 0
        df.replace("--", 0, inplace=True)

        # Create "序号" column starting from 1
        df.sort_values(by="月搜索量", ascending=False, inplace=True)
        df.insert(0, "序号", range(1, len(df) + 1))

        # Extract 自然流量占比 from 贡献流量
        df['自然流量占比'] = df['贡献流量'].str.extract(r':(\d+(?:\.\d+)?)%').fillna(0).astype(float)
        mask = (
            df["出单词来源"].str.contains("热搜关键词", na=False) &
            (df["月搜索量"].astype(float) > 5000) &
            (df["流量占比(%)"].astype(float) > 1)
        )
        # print("Mask: ")
        # print(mask)

        # Filter out unnecessary columns
        required_columns = [
            "序号", "关键词", "出单词来源", "月排名", "月搜索量", "周搜索排名变化", "旺季",
            "曝光形式",	"流量占比(%)",	"点击垄断性(%)", "转化垄断性(%)", "自然流量占比"]
        # Ensure only existing columns are selected (avoid errors if some columns are missing)
        available_columns = [col for col in required_columns if col in df.columns]
        df = df[available_columns]  

        # Print the word frequency
        word_freq = count_word_freq(df, "关键词")
        print("Word frequency: ")
        print(word_freq)
        freq_df = pd.DataFrame(word_freq.items(), columns=["word", "count"]).sort_values(by="count", ascending=False)
        freq_df.to_excel(writer, sheet_name="词频", index=False)

        # Write to the output Excel file with ASIN as the sheet name
        df.to_excel(writer, sheet_name=asin, index=False)

        # Save the final combined file
        writer.close()
        print(f"Processed file saved to: {out}")
        # print(f"Combined file saved to: {output_file}")

        # Open the saved file using openpyxl
        wb = load_workbook(out)
        ws = wb[asin]

        # Map column names to Excel column indexes
        header = [cell.value for cell in ws[1]]
        col_index = {name: idx + 1 for idx, name in enumerate(header)}

        # Define the yellow fill
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Loop through DataFrame rows and corresponding Excel rows
        for i, highlight in enumerate(mask, start=2):  # start=2 because row 1 is header
            if highlight:
                # Color the '关键词' column in the Excel sheet
                ws.cell(row=i, column=col_index["关键词"]).fill = yellow_fill

        # Save changes
        wb.save(out)
        post_process(out)