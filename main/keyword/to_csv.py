import pandas as pd
from playwright.sync_api import sync_playwright
from collections import Counter
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from collections import Counter

ASIN = 'B0DNZ535C6'
product_site = f'https://www.amazon.com/dp/{ASIN}?th=1'

def to_csv():
    # Load the Excel file
    excel_file = "./comments.xlsx"  # Replace with your file name
    df = pd.read_excel(excel_file)
    # Save as CSV
    csv_file = "./output.csv"  # Replace with your desired output file name
    df.to_csv(csv_file, index=False)
    print(f"Conversion complete. CSV saved as {csv_file}")


def test():
    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page()
        page.goto("https://www.amazon.com/sp?ie=UTF8&seller=AF9NNE4MR5EHT&asin=B07CYWKJRY&ref_=dp_merchant_link&isAmazonFulfilled=1")

        # Run JavaScript inside the browser
        content = page.evaluate("""() => {
            return document.querySelector("#page-section-detail-seller-info").innerText;
        }""")
        print(content)
        browser.close()

'''
Count the frequency of each word in the given column
'''
def count_word_freq(df, col_name):
    all_words = []
    for phrase in df[col_name].dropna():
        words = phrase.strip().split()
        all_words.extend(words)
    return Counter(all_words)

# Directory paths
RAW_DIR = "../../raw/keyword"
RESULT_DIR = "../../result/keyword"

# Prompt user whether they want top 400 data
dir_name = input(f"请输入你想要处理的文件夹名称 (注：请先将需要处理的文件合并在一个文件夹中): ").strip()
input_folder = f"{RAW_DIR}/{dir_name}"
output_file = os.path.join(RESULT_DIR, dir_name + "_combined.xlsx")

if __name__ == "__main__":
    input_file = input("Enter the Excel file to process: ")
    to_csv()
    test()