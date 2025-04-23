import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from copy import copy
from openpyxl.utils import get_column_letter

input_file = input("请输入你要处理的文件的路径: ")
title_row = int(input("请输入标题所在的行数 (比如，从上至下第一行，请输入 “1”): "))

# === Step 1: 读取excel ===
df = pd.read_excel(input_file, header=title_row - 1)
df.replace("--", 0, inplace=True)
print("读取成功！")

# === Step 2: 处理 '贡献流量' 数据 ===
def extract_natural_traffic(text):
    try:
        return float(text.split('自然流量:')[1].split('%')[0])
    except:
        return None

if '贡献流量' in df.columns:
    df['自然流量'] = df['贡献流量'].astype(str).apply(extract_natural_traffic)

# === Step 3: 处理 '广告竞争' 数据 ===
def extract_ad_competition(text):
    try:
        return float(text.split('（')[1].split('）')[0])
    except:
        return None

if '广告位竞争' in df.columns:
    df['广告竞争指数'] = df['广告位竞争'].astype(str).apply(extract_ad_competition)

# === Step 4: 只留下想要的数据列 ===
final_columns = ['关键词', '流量来源', '贡献流量', '自然流量', '流量占比(%)', 
       '周搜索排名', '周搜索排名变化', '月搜索量', '广告位竞争', '广告竞争指数', '旺季', '近3个月(%)',
       '近6个月(%)', '近12个月(%)', '词搜索转换比(%)', '90天购买量', 'cpc精准竞价($)',
       'cpc最低竞价($)', 'cpc最高竞价($)', '点击垄断性(%)', '转化垄断性(%)', '点击占比', '转化占比']


print(df.columns)
print("==========")

# Filter and reorder (ignore missing columns silently)
df = df[[col for col in final_columns if col in df.columns]]
print(df.columns)
print("==========")

# === Step 5: 把提取出来的数据存在一个新的表里 ===
temp_file = input_file.replace(".xlsx", "_temp.xlsx")
df.to_excel(temp_file, index=False)


# === Step 6: 染色 ===
wb = load_workbook(temp_file)
ws = wb.active

# 哪一列，什么情况下，染什么色
colors = {
    '流量占比(%)': (lambda v: v >= 2, '9BC2E6'),
    '自然流量': (lambda v: v >= 60, 'F8CBAD'),
    '月搜索量': (lambda v: v >= 5000, 'DAE3F3'),
    '词搜索转换比(%)': (lambda v: 'FFE699' if v >= 5 else 'B4BBC3' if v >= 3 else 'FFFFFF'), 
    'cpc精准竞价($)': (lambda v: v <= 1, 'A9CE91'),
    '广告竞争指数': (lambda v: v <= 3, 'C9C9C9'),
}


# 计算哪一个标题是哪一列
header_map = {cell.value: cell.column for cell in ws[1]}

# 将所有出现在 colors 中的列先转换为数值类型
# 设置目标列为数值格式（默认两位小数）
for header in colors.keys():
    print("Color header:")
    print(header)
    col_idx = header_map.get(header)
    if not col_idx:
        print("not  col_idx")
        continue

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[col_idx - 1]
        try:
            # 把值转成 float（防止仍为字符串）
            if cell.value is not None:
                cell.value = float(str(cell.value).replace('%', '').replace(',', '').strip())
                cell.number_format = '0.00'  # 设置为数值格式
        except:
            continue

# 从第二行开始，一行一行处理数据
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for header, condition_info in colors.items():
        col_idx = header_map.get(header)
        cell = row[col_idx - 1]
        if header == '词搜索转换比(%)':
            try:
                color_code = condition_info(cell.value)
                if cell.value and color_code and color_code.upper() != 'FFFFFF':
                    cell.fill = PatternFill(start_color=color_code, fill_type="solid")
            except:
                continue
        else:
            try: 
                if cell.value is not None and condition_info[0](cell.value):
                    cell.fill = PatternFill(start_color=condition_info[1], fill_type="solid")
            except:
                continue


# === Step 7: 染色后的统计 ===
highlighted_rows = []  # 用于存储染色超过3格的行号
highlighted_counts = {}  # 行号 => 染色格数

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    count = 0
    for cell in row:
        fill = cell.fill
        if fill.fill_type == 'solid' and fill.start_color.rgb not in ['00000000', 'FFFFFFFF', None]:
            count += 1
    highlighted_counts[row_idx] = count
    if count > 3:
        highlighted_rows.append(row_idx)

# 创建新工作表
if '高亮词' in wb.sheetnames:
    del wb['高亮词']
highlight_ws = wb.create_sheet("高亮词")

# 复制标题
for col_idx, cell in enumerate(ws[1], start=1):
    highlight_ws.cell(row=1, column=col_idx).value = cell.value

# # 复制被染色超过3格的行
# for new_row_idx, original_row_idx in enumerate(highlighted_rows, start=2):
#     for col_idx, cell in enumerate(ws[original_row_idx], start=1):
#         highlight_ws.cell(row=new_row_idx, column=col_idx).value = cell.value

# 复制被染色超过3格的行，包括颜色
for new_row_idx, original_row_idx in enumerate(highlighted_rows, start=2):
    for col_idx, cell in enumerate(ws[original_row_idx], start=1):
        new_cell = highlight_ws.cell(row=new_row_idx, column=col_idx)
        new_cell.value = cell.value
        new_cell.fill = copy(cell.fill)

# === Step 8 最后突出标题行的颜色，加上具体的指标 ===
# 设置标题颜色
title_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
title_font = Font(color='FFFFFF', bold=True)

for cell in highlight_ws[1]:
    cell.fill = title_fill
    cell.font = title_font

# 打开指标模板
other_wb = load_workbook(filename="..\..\反查关键词模板.xlsx")
other_ws = other_wb["高亮词"]

# Insert new row at the top of highlight_ws
highlight_ws.insert_rows(1)

# Copy values and formatting from the first row of the source worksheet
for col_idx, cell in enumerate(other_ws[1], start=1):
    target_cell = highlight_ws.cell(row=1, column=col_idx)
    target_cell.value = cell.value
    target_cell.font = copy(cell.font)
    target_cell.fill = copy(cell.fill)

# 设置列宽
for col in highlight_ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
        except:
            pass
    adjusted_width = max_length + 2
    highlight_ws.column_dimensions[col_letter].width = adjusted_width

# Save final file
output_file = input_file.replace(".xlsx", "_formatted.xlsx")
wb.save(output_file)
print(f"Done! Saved as: {output_file}")
